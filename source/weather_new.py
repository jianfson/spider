#!/usr/bin/env python
# coding=utf-8
# agri.gov.cn_amf_client.py
# http://jgsb.agri.gov.cn/flexapps/hqApp.swf数据抓取

import urllib2
import uuid
import pyamf
import json
import os
from openpyxl import Workbook
from openpyxl import load_workbook 
from pyamf import remoting
from pyamf.flex import messaging

def weather_new_func( par ):
    
    # 提交请求
    url = 'http://cdqxapp.cdtq.gov.cn/appcdqx/liverange/type/dryBulTemp/stationid/all/hour/'+''
    req = urllib2.Request(url, headers={'Content-Type': 'application/json'})
    print req
    # 解析返回数据
    opener = urllib2.urlopen(req)
    res = opener.read()
    print "----------------------------------"
    excel_data=json.loads(res)
    print excel_data[0]['dryBulTemp']

    # 提交请求
    url2_1 = 'http://cdqxapp.cdtq.gov.cn/appcdqx/liverange/type/precipitation/stationid/all/hour/'+'1'
    req2_1 = urllib2.Request(url2_1, headers={'Content-Type': 'application/json'})
    print req2_1
    opener2_1 = urllib2.urlopen(req2_1)
    res2_1 = opener2_1.read()
    excel_data2_1=json.loads(res2_1)

    url2_3 = 'http://cdqxapp.cdtq.gov.cn/appcdqx/liverange/type/precipitation/stationid/all/hour/'+'3'
    req2_3 = urllib2.Request(url2_3, headers={'Content-Type': 'application/json'})
    print req2_3
    opener2_3 = urllib2.urlopen(req2_3)
    res2_3 = opener2_3.read()
    excel_data2_3=json.loads(res2_3)

    url2_6 = 'http://cdqxapp.cdtq.gov.cn/appcdqx/liverange/type/precipitation/stationid/all/hour/'+'6'
    req2_6 = urllib2.Request(url2_6, headers={'Content-Type': 'application/json'})
    print req2_6
    opener2_6 = urllib2.urlopen(req2_6)
    res2_6 = opener2_6.read()
    excel_data2_6=json.loads(res2_6)

    url2_12 = 'http://cdqxapp.cdtq.gov.cn/appcdqx/liverange/type/precipitation/stationid/all/hour/'+'12'
    req2_12 = urllib2.Request(url2_12, headers={'Content-Type': 'application/json'})
    print req2_12
    opener2_12 = urllib2.urlopen(req2_12)
    res2_12 = opener2_12.read()
    excel_data2_12=json.loads(res2_12)

    url2_24 = 'http://cdqxapp.cdtq.gov.cn/appcdqx/liverange/type/precipitation/stationid/all/hour/'+'24'
    req2_24 = urllib2.Request(url2_24, headers={'Content-Type': 'application/json'})
    print req2_24
    opener2_24 = urllib2.urlopen(req2_24)
    res2_24 = opener2_24.read()
    excel_data2_24=json.loads(res2_24)

     # 提交请求
    url3 = 'http://cdqxapp.cdtq.gov.cn/appcdqx/liverange/type/relHumidity/stationid/all/hour/'+''
    req3 = urllib2.Request(url3, headers={'Content-Type': 'application/json'})
    print req3
    opener3 = urllib2.urlopen(req3)
    res3 = opener3.read()
    excel_data3=json.loads(res3)


    #print excel_data
    excel_name = "weather_new.xlsx"
    i=0
    while i<len(excel_data):
        sheet_name = excel_data[i]['city']
        print sheet_name
        line = []
        line.append(excel_data[i]['observe_time'])
        line.append(excel_data[i]['dryBulTemp'])
        line.append(excel_data2_1[i]['presum'])
        line.append(excel_data2_3[i]['presum'])
        line.append(excel_data2_6[i]['presum'])
        line.append(excel_data2_12[i]['presum'])
        line.append(excel_data2_24[i]['presum'])
        line.append(excel_data3[i]['relHumidity'])
        p1 = os.path.exists(excel_name)
        if p1:
            wb = load_workbook(excel_name)
            exist_sheet = 0
            for sheet in wb:
                if sheet.title == sheet_name:
                    exist_sheet = 1
                    break
            if exist_sheet == 1:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet()
                ws.title = sheet_name
                line_name = []
                line_name.append('观测时间')
                line_name.append('气温(℃)')
                line_name.append('1小时降水(mm)')
                line_name.append('3小时降水(mm)')
                line_name.append('6小时降水(mm)')
                line_name.append('12小时降水(mm)')
                line_name.append('24小时降水(mm)')
                line_name.append('相对湿度(%)')
                ws.append(line_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            line_name = []
	    line_name.append('观测时间')
	    line_name.append('气温(℃)')
	    line_name.append('1小时降水(mm)')
	    line_name.append('3小时降水(mm)')
            line_name.append('6小时降水(mm)')
	    line_name.append('12小时降水(mm)')
	    line_name.append('24小时降水(mm)')
	    line_name.append('相对湿度(%)')
            ws.append(line_name)
        ws.append(line)
        wb.save(excel_name)
            #print i, excel_data[i]['stationname'], '\n'
        i=i+1
