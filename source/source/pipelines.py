# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from source.items import SourceItem


class SourcePipeline(object):
    def process_item(self, item, spider):
        #p1 = os.path.exists("temp.xlsx")
        #if p1:
        #    print "exist"
        #else:
        #    print "not exist"
        #test_wb = load_workbook("temp.xlsx")
        #self.wb = Workbook()
        #self.ws = self.wb.active
        if isinstance(item, SourceItem):
            ## 在此可进行文件写入、数据库写入等操作
            data = item["data"]
            i=1
            #print len(data)
            line_len = len(data[0])
            name = data[i][1]+".xlsx"
            while i<len(data):
                if data[i][1] != "":
                    name = data[i][1]+".xlsx"
                sheet_name = data[i][2]+"_"+data[i][3]
                p1 = os.path.exists(name)
                if p1:
                    self.wb = load_workbook(name)
                    exist_sheet = 0
                    for sheet in self.wb:
                        if sheet.title == sheet_name:
                            exist_sheet = 1
                            break
                    if exist_sheet == 1:
                        self.ws = self.wb[sheet_name]
                    else:
                        self.ws = self.wb.create_sheet()
                        self.ws.title = sheet_name
                        self.ws.append(data[0][4:line_len])
                else:
                    self.wb = Workbook()
                    #默认插在工作簿末尾
                    #self.ws = self.wb.create_sheet()
                    self.ws = self.wb.active
                    self.ws.title = sheet_name
                    self.ws.append(data[0][4:line_len])
                self.ws.append(data[i][4:line_len])
                self.wb.save(name)
                i=i+1
            #pass
        return item
