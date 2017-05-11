#!/usr/bin/env python
# coding=utf-8
# agri.gov.cn_amf_client.py
# http://jgsb.agri.gov.cn/flexapps/hqApp.swf数据抓取

import urllib2
import uuid
import pyamf
import json
import os
from openpyxl import Workbook
from openpyxl import load_workbook 
from pyamf import remoting
from pyamf.flex import messaging


# 构造flex.messaging.messages.RemotingMessage消息
msg = messaging.RemotingMessage(messageId=str(uuid.uuid1()).upper(),
                                clientId=str(uuid.uuid1()).upper(),
                                operation='queryAutoStationByTime',
                                destination='RealtimeDataQuery',
                                timeToLive=0,
                                timestamp=0)
# 第一个是查询参数，第二个是页数，第三个是控制每页显示的数量（默认每页只显示15条）
msg.body = ['2017-05-11 12:00:00','null', None]
msg.headers['DSEndpoint'] = None
msg.headers['DSId'] = str(uuid.uuid1()).upper()
# 按AMF协议编码数据
req = remoting.Request('null', body=(msg,))
env = remoting.Envelope(amfVersion=pyamf.AMF3)
env.bodies = [('/1', req)]
data = bytes(remoting.encode(env).read())

# 提交请求
url = 'http://resource.cdtq.gov.cn:8080/CDPubServiceWebSite/messagebroker/amf'
req = urllib2.Request(url, data, headers={'Content-Type': 'application/x-amf'})
print req
# 解析返回数据
opener = urllib2.build_opener()
print opener

# 解码AMF协议返回的数据
resp = remoting.decode(opener.open(req).read())
excel = resp.bodies[0][1].body.body
#print excel
jo=json.loads(excel)
#print jo['weather']
excel_data = jo['weather']
print excel_data[0]
excel_name = "weather.xlsx"
i=0
while i<len(excel_data):
    sheet_name = excel_data[i]['stationname']
    line = []
    line.append(excel_data[i]['aa'])
    line.append(excel_data[i]['maxtemp'])
    line.append(excel_data[i]['timeminpsta'])
    line.append(excel_data[i]['mintemp'])
    line.append(excel_data[i]['lon'])
    line.append(excel_data[i]['stationpress'])
    line.append(excel_data[i]['exmaxwindv'])
    line.append(excel_data[i]['lat'])
    line.append(excel_data[i]['stationcode'])
    line.append(excel_data[i]['type'])
    line.append(excel_data[i]['windvelocity'])
    line.append(excel_data[i]['relhumidity'])
    line.append(excel_data[i]['winddirect'])
    line.append(excel_data[i]['drybultemp'])
    line.append(excel_data[i]['precipitation'])
    line.append(excel_data[i]['exmaxwindd'])
    p1 = os.path.exists(excel_name)
    if p1:
        wb = load_workbook(excel_name)
        exist_sheet = 0
        for sheet in wb:
            if sheet.title == sheet_name:
                exist_sheet = 1
                break
        if exist_sheet == 1:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet()
            ws.title = sheet_name
            line_name = []
            line_name.append('time')
            line_name.append('maxtemp')
            line_name.append('timeminpsta')
            line_name.append('mintemp')
            line_name.append('lon')
            line_name.append('stationpress')
            line_name.append('exmaxwindv')
            line_name.append('lat')
            line_name.append('stationcode')
            line_name.append('type')
            line_name.append('windvelocity')
            line_name.append('relhumidity')
            line_name.append('winddirect')
            line_name.append('drybultemp')
            line_name.append('precipitation')
            line_name.append('exmaxwindd')
            ws.append(line_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        line_name = []
        line_name.append('time')
        line_name.append('maxtemp')
        line_name.append('timeminpsta') 
        line_name.append('mintemp') 
        line_name.append('lon')
        line_name.append('stationpress')
        line_name.append('exmaxwindv')
        line_name.append('lat')
        line_name.append('stationcode')
        line_name.append('type')
        line_name.append('windvelocity')
        line_name.append('relhumidity')
        line_name.append('winddirect')
        line_name.append('drybultemp')
        line_name.append('precipitation') 
        line_name.append('exmaxwindd')
        ws.append(line_name)
    ws.append(line)
    wb.save(excel_name)
    #print i, excel_data[i]['stationname'], '\n'
    i=i+1

#for stationname in excel_data:
#    print excel_data['stationname']
#          record['marketName'], \
#          record['maxPrice'], \
#          record['minPrice'], \
#          record['averagePrice'], \
#          record['producAdd'] and record['producAdd'], \
#          record['reportMan']<br>
